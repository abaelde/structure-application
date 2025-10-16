"""
Excel Program Manager - Unified import/export for Excel program files

This module provides a unified interface for loading and saving Program objects
to/from Excel files, ensuring consistency between import and export operations.
"""

import pandas as pd
import numpy as np
from typing import Dict, Any, List, Optional
from pathlib import Path

from src.domain import (
    DIMENSIONS,
    SHEETS,
    PROGRAM_COLS,
    STRUCTURE_COLS,
    Program,
    Structure,
)
from src.domain.dimension_mapping import DIMENSION_COLUMN_MAPPING
from src.domain.models import Condition

# Séparateur pour les dimensions multiples (cohérent partout)
MULTI_VALUE_SEPARATOR = ";"


class ExcelProgramManager:
    """
    Unified manager for Excel program import/export operations.
    
    This class ensures consistency between loading and saving operations by:
    - Using the same data transformation logic for both directions
    - Centralizing Excel column definitions and mappings
    - Providing a single source of truth for Excel format
    """
    
    def __init__(self, source: Optional[str] = None):
        self.source = source
        self.program = None
        self.dimension_columns = []
        
        if source:
            self.load_program()
    
    def load_program(self, source: Optional[str] = None) -> Program:
        """Load a program from Excel file"""
        if source:
            self.source = source
            
        if not self.source:
            raise ValueError("No source file specified")
            
        # Load Excel sheets
        program_df = pd.read_excel(self.source, sheet_name=SHEETS.PROGRAM)
        structures_df = pd.read_excel(self.source, sheet_name=SHEETS.STRUCTURES)
        conditions_df = pd.read_excel(self.source, sheet_name=SHEETS.conditions)
        
        # Process underwriting department
        program_uw_dept = self._convert_pandas_to_native(
            program_df.iloc[0].get(PROGRAM_COLS.UNDERWRITING_DEPARTMENT)
        )
        
        # Process boolean columns
        conditions_df = self._process_boolean_columns(conditions_df, program_uw_dept)
        
        # Determine dimension columns
        self.dimension_columns = self._get_dimension_candidates(conditions_df)
        
        # Parse multi-value dimensions
        for col in self.dimension_columns:
            conditions_df[col] = conditions_df[col].map(self._split_to_list_strict, na_action='ignore')
        
        # Convert to native Python types
        program_name = self._convert_pandas_to_native(program_df.iloc[0][PROGRAM_COLS.TITLE])
        structures_data = self._dataframe_to_dict_list(structures_df)
        conditions_data = self._dataframe_to_dict_list(conditions_df)
        
        # Build structures and conditions
        structures = self._build_structures(structures_data, conditions_data)
        
        # Create program
        self.program = Program(
            name=program_name,
            structures=structures,
            dimension_columns=self.dimension_columns,
            underwriting_department=program_uw_dept,
        )
        
        return self.program
    
    def save_program(self, program: Program, output_path: str, min_width: int = 10, max_width: int = 50) -> None:
        """Save a program to Excel file"""
        # Create Excel data structures
        excel_data = self._program_to_excel_data(program)
        
        # Write to Excel
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            excel_data["program"].to_excel(writer, sheet_name=SHEETS.PROGRAM, index=False)
            excel_data["structures"].to_excel(writer, sheet_name=SHEETS.STRUCTURES, index=False)
            excel_data["conditions"].to_excel(writer, sheet_name=SHEETS.conditions, index=False)
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(output_path, min_width=min_width, max_width=max_width)
    
    def get_program(self) -> Program:
        """Get the loaded program"""
        if not self.program:
            raise ValueError("No program loaded. Call load_program() first.")
        return self.program
    
    # Private helper methods for data transformation
    
    def _split_to_list_strict(self, cell) -> list[str] | None:
        """Parse multi-value cells (same logic for import)"""
        if pd.isna(cell):
            return None
        tokens = [t.strip() for t in str(cell).split(MULTI_VALUE_SEPARATOR)]
        tokens = [t for t in tokens if t]  # Remove empty
        if not tokens:
            return None
        return tokens
    
    def _convert_list_to_excel_string(self, value) -> str | None:
        """Convert list values to semicolon-separated strings (for export)"""
        if isinstance(value, list):
            return MULTI_VALUE_SEPARATOR.join(map(str, value))
        return value
    
    def _convert_pandas_to_native(self, value):
        """Convert pandas values to native Python types"""
        if isinstance(value, list):
            return value
        if pd.isna(value):
            return None
        return value
    
    def _get_dimension_candidates(self, conditions_df: pd.DataFrame) -> list[str]:
        """Get dimension columns (same logic for both import/export)"""
        program_dims = set(DIMENSIONS) | set(DIMENSION_COLUMN_MAPPING.keys())
        program_dims -= {"INCLUDES_HULL", "INCLUDES_LIABILITY"}  # Boolean flags
        return [c for c in program_dims if c in conditions_df.columns]
    
    def _dataframe_to_dict_list(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Convert DataFrame to list of dicts"""
        records = df.to_dict("records")
        return [
            {key: self._convert_pandas_to_native(value) for key, value in record.items()}
            for record in records
        ]
    
    def _process_boolean_columns(self, conditions_df: pd.DataFrame, underwriting_department: str) -> pd.DataFrame:
        """Process boolean columns (same logic as original loader)"""
        boolean_columns = ["INCLUDES_HULL", "INCLUDES_LIABILITY"]
        
        for col in boolean_columns:
            if col in conditions_df.columns:
                conditions_df[col] = conditions_df[col].map(self._to_boolean, na_action='ignore')
        
        if underwriting_department and underwriting_department.lower() == "aviation":
            for col in boolean_columns:
                if col not in conditions_df.columns:
                    raise ValueError(
                        f"Column {col} is required for Aviation programs. "
                        f"Please add {col} column to the conditions sheet with explicit TRUE/FALSE values."
                    )
                
                null_values = conditions_df[col].isna().sum()
                if null_values > 0:
                    raise ValueError(
                        f"Column {col} has {null_values} null/empty values. "
                        f"For Aviation programs, all conditions must have explicit TRUE or FALSE values for {col}."
                    )
        
        return conditions_df
    
    @staticmethod
    def _to_boolean(value):
        """Convert value to boolean (same logic as original loader)"""
        if pd.isna(value):
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            value_lower = value.lower().strip()
            if value_lower in ("true", "1", "yes", "y"):
                return True
            if value_lower in ("false", "0", "no", "n"):
                return False
            return None
        if isinstance(value, (int, float)):
            return bool(value)
        return None
    
    def _build_structures(self, structures_data: List[Dict[str, Any]], conditions_data: List[Dict[str, Any]]) -> List[Structure]:
        """Build Structure objects from data"""
        # Group conditions by structure
        conditions_by_structure = {}
        for condition in conditions_data:
            structure_key = condition.get(STRUCTURE_COLS.INSPER_ID)
            if structure_key is None:
                raise ValueError(
                    f"Structure key (INSPER_ID_PRE) is mandatory for all conditions. "
                    f"Found condition without structure key: {condition}"
                )
            
            if structure_key not in conditions_by_structure:
                conditions_by_structure[structure_key] = []
            conditions_by_structure[structure_key].append(condition)
        
        # Build structures
        structures = []
        for structure_dict in structures_data:
            structure_key = structure_dict.get(STRUCTURE_COLS.INSPER_ID)
            if structure_key is None:
                raise ValueError(
                    f"Structure key (INSPER_ID_PRE) is mandatory for all structures. "
                    f"Found structure without structure key: {structure_dict}"
                )
            
            conditions_data_for_structure = conditions_by_structure.get(structure_key, [])
            structures.append(
                Structure.from_row(structure_dict, conditions_data_for_structure, STRUCTURE_COLS)
            )
        
        return structures
    
    def _program_to_excel_data(self, program: Program) -> Dict[str, pd.DataFrame]:
        """Convert Program to Excel DataFrames (unified export logic)"""
        reprog_id = 1
        
        # Program sheet
        program_df = pd.DataFrame({
            "REPROG_ID_PRE": [reprog_id],
            "REPROG_TITLE": [program.name],
            "CED_ID_PRE": [None],
            "CED_NAME_PRE": [None],
            "REPROG_ACTIVE_IND": [True],
            "REPROG_COMMENT": [None],
            "REPROG_UW_DEPARTMENT_CD": [None],
            "REPROG_UW_DEPARTMENT_NAME": [None],
            "REPROG_UW_DEPARTMENT_LOB_CD": [program.underwriting_department],
            "REPROG_UW_DEPARTMENT_LOB_NAME": [program.underwriting_department.title() if program.underwriting_department else None],
            "BUSPAR_CED_REG_CLASS_CD": [None],
            "BUSPAR_CED_REG_CLASS_NAME": [None],
            "REPROG_MAIN_CURRENCY_CD": [None],
            "REPROG_MANAGEMENT_REPORTING_LOB_CD": [None],
        })
        
        # Initialize data structures
        structures_data = self._init_structures_data()
        conditions_data = self._init_conditions_data()
        
        insper_id = 1
        condition_id = 1
        
        # Process structures and conditions
        for structure in program.structures:
            # Add structure data
            self._add_structure_data(structures_data, structure, insper_id, reprog_id)
            
            # Add conditions data
            for condition in structure.conditions:
                self._add_condition_data(conditions_data, condition, condition_id, insper_id, reprog_id)
                condition_id += 1
            
            insper_id += 1
        
        return {
            "program": program_df,
            "structures": pd.DataFrame(structures_data),
            "conditions": pd.DataFrame(conditions_data)
        }
    
    def _init_structures_data(self) -> Dict[str, List]:
        """Initialize structures data structure"""
        return {
            "INSPER_ID_PRE": [],
            "BUSINESS_ID_PRE": [],
            "TYPE_OF_PARTICIPATION_CD": [],
            "TYPE_OF_INSURED_PERIOD_CD": [],
            "ACTIVE_FLAG_CD": [],
            "INSPER_EFFECTIVE_DATE": [],
            "INSPER_EXPIRY_DATE": [],
            "REPROG_ID_PRE": [],
            "BUSINESS_TITLE": [],
            "INSPER_LAYER_NO": [],
            "INSPER_MAIN_CURRENCY_CD": [],
            "INSPER_UW_YEAR": [],
            "INSPER_CONTRACT_ORDER": [],
            "INSPER_PREDECESSOR_TITLE": [],
            "INSPER_CONTRACT_FORM_CD_SLAV": [],
            "INSPER_CONTRACT_LODRA_CD_SLAV": [],
            "INSPER_CONTRACT_COVERAGE_CD_SLAV": [],
            "INSPER_CLAIM_BASIS_CD": [],
            "INSPER_LODRA_CD_SLAV": [],
            "INSPER_LOD_TO_RA_DATE_SLAV": [],
            "INSPER_COMMENT": [],
        }
    
    def _init_conditions_data(self) -> Dict[str, List]:
        """Initialize conditions data structure"""
        return {
            "BUSCL_ID_PRE": [],
            "REPROG_ID_PRE": [],
            "CED_ID_PRE": [],
            "BUSINESS_ID_PRE": [],
            "INSPER_ID_PRE": [],
            "BUSCL_EXCLUDE_CD": [],
            "BUSCL_ENTITY_NAME_CED": [],
            "POL_RISK_NAME_CED": [],
            "BUSCL_COUNTRY_CD": [],
            "BUSCL_COUNTRY": [],
            "BUSCL_REGION": [],
            "BUSCL_CLASS_OF_BUSINESS_1": [],
            "BUSCL_CLASS_OF_BUSINESS_2": [],
            "BUSCL_CLASS_OF_BUSINESS_3": [],
            "BUSCL_LIMIT_CURRENCY_CD": [],
            "AAD_100": [],
            "LIMIT_100": [],
            "LIMIT_FLOATER_100": [],
            "ATTACHMENT_POINT_100": [],
            "OLW_100": [],
            "LIMIT_OCCURRENCE_100": [],
            "LIMIT_AGG_100": [],
            "CESSION_PCT": [],
            "RETENTION_PCT": [],
            "SUPI_100": [],
            "BUSCL_PREMIUM_CURRENCY_CD": [],
            "BUSCL_PREMIUM_GROSS_NET_CD": [],
            "PREMIUM_RATE_PCT": [],
            "PREMIUM_DEPOSIT_100": [],
            "PREMIUM_MIN_100": [],
            "BUSCL_LIABILITY_1_LINE_100": [],
            "MAX_COVER_PCT": [],
            "MIN_EXCESS_PCT": [],
            "SIGNED_SHARE_PCT": [],
            "AVERAGE_LINE_SLAV_CED": [],
            "PML_DEFAULT_PCT": [],
            "LIMIT_EVENT": [],
            "NO_OF_REINSTATEMENTS": [],
            "INCLUDES_HULL": [],
            "INCLUDES_LIABILITY": [],
        }
    
    def _add_structure_data(self, structures_data: Dict[str, List], structure: Structure, insper_id: int, reprog_id: int):
        """Add structure data to the structures_data dict"""
        structures_data["INSPER_ID_PRE"].append(insper_id)
        structures_data["BUSINESS_ID_PRE"].append(None)
        structures_data["TYPE_OF_PARTICIPATION_CD"].append(structure.type_of_participation)
        structures_data["TYPE_OF_INSURED_PERIOD_CD"].append(None)
        structures_data["ACTIVE_FLAG_CD"].append(True)
        structures_data["INSPER_EFFECTIVE_DATE"].append(structure.inception_date)
        structures_data["INSPER_EXPIRY_DATE"].append(structure.expiry_date)
        structures_data["REPROG_ID_PRE"].append(reprog_id)
        structures_data["BUSINESS_TITLE"].append(structure.structure_name)
        structures_data["INSPER_LAYER_NO"].append(None)
        structures_data["INSPER_MAIN_CURRENCY_CD"].append(None)
        structures_data["INSPER_UW_YEAR"].append(None)
        structures_data["INSPER_CONTRACT_ORDER"].append(None)
        structures_data["INSPER_PREDECESSOR_TITLE"].append(structure.predecessor_title)
        structures_data["INSPER_CONTRACT_FORM_CD_SLAV"].append(None)
        structures_data["INSPER_CONTRACT_LODRA_CD_SLAV"].append(None)
        structures_data["INSPER_CONTRACT_COVERAGE_CD_SLAV"].append(None)
        structures_data["INSPER_CLAIM_BASIS_CD"].append(structure.claim_basis)
        structures_data["INSPER_LODRA_CD_SLAV"].append(None)
        structures_data["INSPER_LOD_TO_RA_DATE_SLAV"].append(None)
        structures_data["INSPER_COMMENT"].append(None)
    
    def _add_condition_data(self, conditions_data: Dict[str, List], condition: Condition, condition_id: int, insper_id: int, reprog_id: int):
        """Add condition data to the conditions_data dict"""
        condition_dict = condition.to_dict()
        
        conditions_data["BUSCL_ID_PRE"].append(condition_id)
        conditions_data["REPROG_ID_PRE"].append(reprog_id)
        conditions_data["CED_ID_PRE"].append(None)
        conditions_data["BUSINESS_ID_PRE"].append(None)
        conditions_data["INSPER_ID_PRE"].append(insper_id)
        conditions_data["BUSCL_EXCLUDE_CD"].append(condition_dict.get("BUSCL_EXCLUDE_CD"))
        
        # Convert dimension lists to semicolon-separated strings
        conditions_data["BUSCL_ENTITY_NAME_CED"].append(self._convert_list_to_excel_string(condition_dict.get("BUSCL_ENTITY_NAME_CED")))
        conditions_data["POL_RISK_NAME_CED"].append(self._convert_list_to_excel_string(condition_dict.get("POL_RISK_NAME_CED")))
        conditions_data["BUSCL_COUNTRY_CD"].append(self._convert_list_to_excel_string(condition_dict.get("BUSCL_COUNTRY_CD")))
        conditions_data["BUSCL_COUNTRY"].append(self._convert_list_to_excel_string(condition_dict.get("BUSCL_COUNTRY")))
        conditions_data["BUSCL_REGION"].append(self._convert_list_to_excel_string(condition_dict.get("BUSCL_REGION")))
        conditions_data["BUSCL_CLASS_OF_BUSINESS_1"].append(self._convert_list_to_excel_string(condition_dict.get("BUSCL_CLASS_OF_BUSINESS_1")))
        conditions_data["BUSCL_CLASS_OF_BUSINESS_2"].append(self._convert_list_to_excel_string(condition_dict.get("BUSCL_CLASS_OF_BUSINESS_2")))
        conditions_data["BUSCL_CLASS_OF_BUSINESS_3"].append(self._convert_list_to_excel_string(condition_dict.get("BUSCL_CLASS_OF_BUSINESS_3")))
        conditions_data["BUSCL_LIMIT_CURRENCY_CD"].append(self._convert_list_to_excel_string(condition_dict.get("BUSCL_LIMIT_CURRENCY_CD")))
        
        # Other fields
        conditions_data["AAD_100"].append(condition_dict.get("AAD_100"))
        conditions_data["LIMIT_100"].append(condition_dict.get("LIMIT_100"))
        conditions_data["LIMIT_FLOATER_100"].append(condition_dict.get("LIMIT_FLOATER_100"))
        
        attachment = condition_dict.get("ATTACHMENT_POINT_100")
        conditions_data["ATTACHMENT_POINT_100"].append(attachment if pd.notna(attachment) else np.nan)
        
        conditions_data["OLW_100"].append(condition_dict.get("OLW_100"))
        conditions_data["LIMIT_OCCURRENCE_100"].append(None)
        conditions_data["LIMIT_AGG_100"].append(condition_dict.get("LIMIT_AGG_100"))
        
        cession = condition_dict.get("CESSION_PCT")
        conditions_data["CESSION_PCT"].append(cession if pd.notna(cession) else np.nan)
        
        conditions_data["RETENTION_PCT"].append(condition_dict.get("RETENTION_PCT"))
        conditions_data["SUPI_100"].append(condition_dict.get("SUPI_100"))
        conditions_data["BUSCL_PREMIUM_CURRENCY_CD"].append(condition_dict.get("BUSCL_PREMIUM_CURRENCY_CD"))
        conditions_data["BUSCL_PREMIUM_GROSS_NET_CD"].append(condition_dict.get("BUSCL_PREMIUM_GROSS_NET_CD"))
        conditions_data["PREMIUM_RATE_PCT"].append(condition_dict.get("PREMIUM_RATE_PCT"))
        conditions_data["PREMIUM_DEPOSIT_100"].append(condition_dict.get("PREMIUM_DEPOSIT_100"))
        conditions_data["PREMIUM_MIN_100"].append(condition_dict.get("PREMIUM_MIN_100"))
        conditions_data["BUSCL_LIABILITY_1_LINE_100"].append(condition_dict.get("BUSCL_LIABILITY_1_LINE_100"))
        conditions_data["MAX_COVER_PCT"].append(condition_dict.get("MAX_COVER_PCT"))
        conditions_data["MIN_EXCESS_PCT"].append(condition_dict.get("MIN_EXCESS_PCT"))
        conditions_data["SIGNED_SHARE_PCT"].append(condition_dict.get("SIGNED_SHARE_PCT"))
        conditions_data["AVERAGE_LINE_SLAV_CED"].append(condition_dict.get("AVERAGE_LINE_SLAV_CED"))
        conditions_data["PML_DEFAULT_PCT"].append(condition_dict.get("PML_DEFAULT_PCT"))
        conditions_data["LIMIT_EVENT"].append(condition_dict.get("LIMIT_EVENT"))
        conditions_data["NO_OF_REINSTATEMENTS"].append(condition_dict.get("NO_OF_REINSTATEMENTS"))
        conditions_data["INCLUDES_HULL"].append(condition_dict.get("INCLUDES_HULL"))
        conditions_data["INCLUDES_LIABILITY"].append(condition_dict.get("INCLUDES_LIABILITY"))
    
    def _auto_adjust_column_widths(self, excel_path: str, min_width: int = 10, max_width: int = 50):
        """Auto-adjust column widths in Excel file"""
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        workbook = load_workbook(excel_path)
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, min_width), max_width)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(excel_path)


# Convenience functions for backward compatibility
def load_program_from_excel(file_path: str) -> Program:
    """Load a program from Excel file (convenience function)"""
    manager = ExcelProgramManager(file_path)
    return manager.get_program()


def save_program_to_excel(program: Program, file_path: str, min_width: int = 10, max_width: int = 50) -> None:
    """Save a program to Excel file (convenience function)"""
    manager = ExcelProgramManager()
    manager.save_program(program, file_path, min_width, max_width)
