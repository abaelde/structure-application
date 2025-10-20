# src/io/excel_adapter.py
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from src.domain.constants import SHEETS


class ExcelProgramIO:
    def read(self, source: str):
        program_df = pd.read_excel(source, sheet_name=SHEETS.PROGRAM)
        structures_df = pd.read_excel(source, sheet_name=SHEETS.STRUCTURES)
        conditions_df = pd.read_excel(source, sheet_name=SHEETS.conditions)
        try:
            exclusions_df = pd.read_excel(source, sheet_name=SHEETS.EXCLUSIONS)
        except Exception:
            exclusions_df = pd.DataFrame()
        return program_df, structures_df, conditions_df, exclusions_df

    def write(
        self,
        dest: str,
        program_df,
        structures_df,
        conditions_df,
        exclusions_df,
        *,
        min_width=10,
        max_width=50,
    ):
        with pd.ExcelWriter(dest, engine="openpyxl") as writer:
            program_df.to_excel(writer, sheet_name=SHEETS.PROGRAM, index=False)
            structures_df.to_excel(writer, sheet_name=SHEETS.STRUCTURES, index=False)
            conditions_df.to_excel(writer, sheet_name=SHEETS.conditions, index=False)
            # exclusions sheet may be missing in older programs
            if "exclusions" in locals() or "exclusions_df" in locals():
                try:
                    exclusions_df.to_excel(writer, sheet_name=SHEETS.EXCLUSIONS, index=False)
                except Exception:
                    # if caller didn't pass it, create empty sheet
                    pd.DataFrame().to_excel(writer, sheet_name=SHEETS.EXCLUSIONS, index=False)
        self._auto_adjust_column_widths(dest, min_width, max_width)

    def _auto_adjust_column_widths(
        self, excel_path: str, min_width: int = 10, max_width: int = 50
    ):
        wb = load_workbook(excel_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_len = 0
                col_letter = get_column_letter(column[0].column)
                for cell in column:
                    v = cell.value
                    if v:
                        max_len = max(max_len, len(str(v)))
                ws.column_dimensions[col_letter].width = min(
                    max(max_len + 2, min_width), max_width
                )
        wb.save(excel_path)
