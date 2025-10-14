"""
Excel utilities for program creation

Provides helper functions for Excel file manipulation, including auto-sizing columns.
"""

import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), "..", ".."))

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
from src.domain import PRODUCT, SHEETS


def auto_adjust_column_widths(
    excel_path: str, min_width: int = 10, max_width: int = 50
):
    """
    Automatically adjust column widths in an Excel file based on content

    Args:
        excel_path: Path to the Excel file
        min_width: Minimum column width (default: 10)
        max_width: Maximum column width (default: 50)
    """
    workbook = load_workbook(excel_path)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    # Calculate cell content length
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Set column width with min/max constraints
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(excel_path)


def create_excel_with_auto_width(
    file_path: str, dataframes: dict, min_width: int = 10, max_width: int = 50
):
    """
    Create an Excel file with multiple sheets and auto-adjusted column widths

    Args:
        file_path: Path for the output Excel file
        dataframes: Dictionary with sheet names as keys and DataFrames as values
        min_width: Minimum column width (default: 10)
        max_width: Maximum column width (default: 50)

    Example:
        create_excel_with_auto_width(
            "output.xlsx",
            {
                "program": program_df,
                "structures": structures_df,
                "sections": sections_df
            }
        )
    """
    import pandas as pd

    # Write DataFrames to Excel
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        for sheet_name, df in dataframes.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Auto-adjust column widths
    auto_adjust_column_widths(file_path, min_width=min_width, max_width=max_width)


def program_to_excel(
    program,
    output_path: str,
    min_width: int = 10,
    max_width: int = 50,
):
    """
    Convert a Program object (created with Builders) to an Excel file.
    
    This function automatically handles:
    - Auto-incrementing IDs (REPROG_ID_PRE, INSPER_ID_PRE, BUSCL_ID_PRE)
    - All required Excel fields with proper defaults
    - Proper structure references (INSPER_ID_PRE in sections)
    - Column width auto-adjustment
    
    Args:
        program: Program object created with build_program()
        output_path: Path for the output Excel file
        min_width: Minimum column width (default: 10)
        max_width: Maximum column width (default: 50)
    
    Example:
        from tests.builders import build_quota_share, build_program
        from examples.program_creation.excel_utils import program_to_excel
        
        qs = build_quota_share(name="QS_30", cession_pct=0.30)
        program = build_program(
            name="SINGLE_QUOTA_SHARE_2024",
            structures=[qs],
            underwriting_department="test"
        )
        
        program_to_excel(program, "../programs/single_quota_share.xlsx")
    """
    reprog_id = 1
    
    program_df = pd.DataFrame({
        "REPROG_ID_PRE": [reprog_id],
        "REPROG_TITLE": [program.name],
        "CED_ID_PRE": [None],
        "CED_NAME_PRE": [None],
        "REPROG_ACTIVE_IND": [True],
        "REPROG_COMMENT": [None],
        "REPROG_UW_DEPARTMENT_CD": [None],
        "REPROG_UW_DEPARTMENT_NAME": [None],
        "REPROG_UW_DEPARTMENT_LOB_CD": [program.underwriting_department],
        "REPROG_UW_DEPARTMENT_LOB_NAME": [program.underwriting_department.title() if program.underwriting_department else None],
        "BUSPAR_CED_REG_CLASS_CD": [None],
        "BUSPAR_CED_REG_CLASS_NAME": [None],
        "REPROG_MAIN_CURRENCY_CD": [None],
        "REPROG_MANAGEMENT_REPORTING_LOB_CD": [None],
    })
    
    structures_data = {
        "INSPER_ID_PRE": [],
        "BUSINESS_ID_PRE": [],
        "TYPE_OF_PARTICIPATION_CD": [],
        "TYPE_OF_INSURED_PERIOD_CD": [],
        "ACTIVE_FLAG_CD": [],
        "INSPER_EFFECTIVE_DATE": [],
        "INSPER_EXPIRY_DATE": [],
        "REPROG_ID_PRE": [],
        "BUSINESS_TITLE": [],
        "INSPER_LAYER_NO": [],
        "INSPER_MAIN_CURRENCY_CD": [],
        "INSPER_UW_YEAR": [],
        "INSPER_CONTRACT_ORDER": [],
        "INSPER_PREDECESSOR_TITLE": [],
        "INSPER_CONTRACT_FORM_CD_SLAV": [],
        "INSPER_CONTRACT_LODRA_CD_SLAV": [],
        "INSPER_CONTRACT_COVERAGE_CD_SLAV": [],
        "INSPER_CLAIM_BASIS_CD": [],
        "INSPER_LODRA_CD_SLAV": [],
        "INSPER_LOD_TO_RA_DATE_SLAV": [],
        "INSPER_COMMENT": [],
    }
    
    sections_data = {
        "BUSCL_ID_PRE": [],
        "REPROG_ID_PRE": [],
        "CED_ID_PRE": [],
        "BUSINESS_ID_PRE": [],
        "INSPER_ID_PRE": [],
        "BUSCL_EXCLUDE_CD": [],
        "BUSCL_ENTITY_NAME_CED": [],
        "POL_RISK_NAME_CED": [],
        "BUSCL_COUNTRY_CD": [],
        "BUSCL_COUNTRY": [],
        "BUSCL_REGION": [],
        "BUSCL_CLASS_OF_BUSINESS_1": [],
        "BUSCL_CLASS_OF_BUSINESS_2": [],
        "BUSCL_CLASS_OF_BUSINESS_3": [],
        "BUSCL_LIMIT_CURRENCY_CD": [],
        "AAD_100": [],
        "LIMIT_100": [],
        "LIMIT_FLOATER_100": [],
        "ATTACHMENT_POINT_100": [],
        "OLW_100": [],
        "LIMIT_OCCURRENCE_100": [],
        "LIMIT_AGG_100": [],
        "CESSION_PCT": [],
        "RETENTION_PCT": [],
        "SUPI_100": [],
        "BUSCL_PREMIUM_CURRENCY_CD": [],
        "BUSCL_PREMIUM_GROSS_NET_CD": [],
        "PREMIUM_RATE_PCT": [],
        "PREMIUM_DEPOSIT_100": [],
        "PREMIUM_MIN_100": [],
        "BUSCL_LIABILITY_1_LINE_100": [],
        "MAX_COVER_PCT": [],
        "MIN_EXCESS_PCT": [],
        "SIGNED_SHARE_PCT": [],
        "AVERAGE_LINE_SLAV_CED": [],
        "PML_DEFAULT_PCT": [],
        "LIMIT_EVENT": [],
        "NO_OF_REINSTATEMENTS": [],
    }
    
    insper_id = 1
    section_id = 1
    
    for structure in program.structures:
        structures_data["INSPER_ID_PRE"].append(insper_id)
        structures_data["BUSINESS_ID_PRE"].append(None)
        structures_data["TYPE_OF_PARTICIPATION_CD"].append(structure.type_of_participation)
        structures_data["TYPE_OF_INSURED_PERIOD_CD"].append(None)
        structures_data["ACTIVE_FLAG_CD"].append(True)
        structures_data["INSPER_EFFECTIVE_DATE"].append(structure.inception_date)
        structures_data["INSPER_EXPIRY_DATE"].append(structure.expiry_date)
        structures_data["REPROG_ID_PRE"].append(reprog_id)
        structures_data["BUSINESS_TITLE"].append(structure.structure_name)
        structures_data["INSPER_LAYER_NO"].append(None)
        structures_data["INSPER_MAIN_CURRENCY_CD"].append(None)
        structures_data["INSPER_UW_YEAR"].append(None)
        structures_data["INSPER_CONTRACT_ORDER"].append(None)
        structures_data["INSPER_PREDECESSOR_TITLE"].append(structure.predecessor_title)
        structures_data["INSPER_CONTRACT_FORM_CD_SLAV"].append(None)
        structures_data["INSPER_CONTRACT_LODRA_CD_SLAV"].append(None)
        structures_data["INSPER_CONTRACT_COVERAGE_CD_SLAV"].append(None)
        structures_data["INSPER_CLAIM_BASIS_CD"].append(structure.claim_basis)
        structures_data["INSPER_LODRA_CD_SLAV"].append(None)
        structures_data["INSPER_LOD_TO_RA_DATE_SLAV"].append(None)
        structures_data["INSPER_COMMENT"].append(None)
        
        for section in structure.sections:
            section_dict = section.to_dict()
            
            sections_data["BUSCL_ID_PRE"].append(section_id)
            sections_data["REPROG_ID_PRE"].append(reprog_id)
            sections_data["CED_ID_PRE"].append(None)
            sections_data["BUSINESS_ID_PRE"].append(None)
            sections_data["INSPER_ID_PRE"].append(insper_id)
            sections_data["BUSCL_EXCLUDE_CD"].append(section_dict.get("BUSCL_EXCLUDE_CD"))
            sections_data["BUSCL_ENTITY_NAME_CED"].append(section_dict.get("BUSCL_ENTITY_NAME_CED"))
            sections_data["POL_RISK_NAME_CED"].append(section_dict.get("POL_RISK_NAME_CED"))
            sections_data["BUSCL_COUNTRY_CD"].append(section_dict.get("BUSCL_COUNTRY_CD"))
            sections_data["BUSCL_COUNTRY"].append(section_dict.get("BUSCL_COUNTRY"))
            sections_data["BUSCL_REGION"].append(section_dict.get("BUSCL_REGION"))
            sections_data["BUSCL_CLASS_OF_BUSINESS_1"].append(section_dict.get("BUSCL_CLASS_OF_BUSINESS_1"))
            sections_data["BUSCL_CLASS_OF_BUSINESS_2"].append(section_dict.get("BUSCL_CLASS_OF_BUSINESS_2"))
            sections_data["BUSCL_CLASS_OF_BUSINESS_3"].append(section_dict.get("BUSCL_CLASS_OF_BUSINESS_3"))
            sections_data["BUSCL_LIMIT_CURRENCY_CD"].append(section_dict.get("BUSCL_LIMIT_CURRENCY_CD"))
            sections_data["AAD_100"].append(section_dict.get("AAD_100"))
            sections_data["LIMIT_100"].append(section_dict.get("LIMIT_100"))
            sections_data["LIMIT_FLOATER_100"].append(section_dict.get("LIMIT_FLOATER_100"))
            
            attachment = section_dict.get("ATTACHMENT_POINT_100")
            sections_data["ATTACHMENT_POINT_100"].append(attachment if pd.notna(attachment) else np.nan)
            
            sections_data["OLW_100"].append(section_dict.get("OLW_100"))
            sections_data["LIMIT_OCCURRENCE_100"].append(None)
            sections_data["LIMIT_AGG_100"].append(section_dict.get("LIMIT_AGG_100"))
            
            cession = section_dict.get("CESSION_PCT")
            sections_data["CESSION_PCT"].append(cession if pd.notna(cession) else np.nan)
            
            sections_data["RETENTION_PCT"].append(section_dict.get("RETENTION_PCT"))
            sections_data["SUPI_100"].append(section_dict.get("SUPI_100"))
            sections_data["BUSCL_PREMIUM_CURRENCY_CD"].append(section_dict.get("BUSCL_PREMIUM_CURRENCY_CD"))
            sections_data["BUSCL_PREMIUM_GROSS_NET_CD"].append(section_dict.get("BUSCL_PREMIUM_GROSS_NET_CD"))
            sections_data["PREMIUM_RATE_PCT"].append(section_dict.get("PREMIUM_RATE_PCT"))
            sections_data["PREMIUM_DEPOSIT_100"].append(section_dict.get("PREMIUM_DEPOSIT_100"))
            sections_data["PREMIUM_MIN_100"].append(section_dict.get("PREMIUM_MIN_100"))
            sections_data["BUSCL_LIABILITY_1_LINE_100"].append(section_dict.get("BUSCL_LIABILITY_1_LINE_100"))
            sections_data["MAX_COVER_PCT"].append(section_dict.get("MAX_COVER_PCT"))
            sections_data["MIN_EXCESS_PCT"].append(section_dict.get("MIN_EXCESS_PCT"))
            sections_data["SIGNED_SHARE_PCT"].append(section_dict.get("SIGNED_SHARE_PCT"))
            sections_data["AVERAGE_LINE_SLAV_CED"].append(section_dict.get("AVERAGE_LINE_SLAV_CED"))
            sections_data["PML_DEFAULT_PCT"].append(section_dict.get("PML_DEFAULT_PCT"))
            sections_data["LIMIT_EVENT"].append(section_dict.get("LIMIT_EVENT"))
            sections_data["NO_OF_REINSTATEMENTS"].append(section_dict.get("NO_OF_REINSTATEMENTS"))
            
            section_id += 1
        
        insper_id += 1
    
    structures_df = pd.DataFrame(structures_data)
    sections_df = pd.DataFrame(sections_data)
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        program_df.to_excel(writer, sheet_name=SHEETS.PROGRAM, index=False)
        structures_df.to_excel(writer, sheet_name=SHEETS.STRUCTURES, index=False)
        sections_df.to_excel(writer, sheet_name=SHEETS.SECTIONS, index=False)
    
    auto_adjust_column_widths(output_path, min_width=min_width, max_width=max_width)
